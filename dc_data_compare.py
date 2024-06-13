import pandas as pd
import openpyxl
import openpyxl as px
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog

def compare_excel_files(file1, file2, site_to_compare):
    """
    Compare two Excel files and save the differences to a new Excel file.

    Args:
        file1 (str): Path to the first Excel file.
        file2 (str): Path to the second Excel file.
    """
    # Put the Excel files into DataFrames
    df1 = pd.read_excel(file1, dtype={'Node': str})
    df2 = pd.read_excel(file2, dtype={'Node': str}) 

    # Strip leading/trailing spaces from column names
    df1.columns = df1.columns.str.strip()
    df2.columns = df2.columns.str.strip()

    # Filter df1 to only include rows where the 'Site ID' is 'REDACTED_03'(example)
    df1_filtered = df1[df1['Site ID'].str.strip() == site_to_compare]

    # Convert 'Node' column to string type and ensure all values are numeric strings with leading zeros
    df1_filtered['Node'] = df1_filtered['Node'].apply(lambda x: f'{int(x):03}' if pd.notnull(x) else '000')
    df2['Node'] = df2['Node'].apply(lambda x: f'{int(x):03}' if pd.notnull(x) else '000')

    # Set the 'Physical UHN' column as the index for both DataFrames to align data
    df1_filtered.set_index('Physical UHN', inplace=True)
    df2.set_index('Physical UHN', inplace=True)

    # Make sure index is unique
    df1_filtered = df1_filtered[~df1_filtered.index.duplicated(keep='first')]
    df2 = df2[~df2.index.duplicated(keep='first')]

    # The list of columns to compare
    columns_to_compare = [
        'Status', 'Site ID', 'Building Name', 'Floor Name', 'Room Name', 'Zone Name', 
        'Row Name', 'Rack Name', 'POD Code', 'Material Name', 'Device Role', 'Node', 
        'Serial Number', 'Mac Address', 'Material Start Slot Number', 'Material End Slot Number', 
        'Number of Slots', 'Product Number', 'Material Code', 'Rack Sequence Number', 
        'Logical UHN', 'Cluster Code', 'Server Modified', 'Current SKU Type'
    ]

    # Initialize a list to store differences
    differences = []

    # Compare rows and columns
    for index in df1_filtered.index.union(df2.index):  # union to include all indices from both dataframes
        value1 = df1_filtered.index.get_loc(index) if index in df1_filtered.index else None
        value2 = df2.index.get_loc(index) if index in df2.index else None
        '''potentially garbage code'''
        # Check if 'Physical UHN' is blank in either file
        #if pd.isna(value1) or pd.isna(value2):
            #differences.append({
                #'Physical UHN': index,
                #'Column': 'Physical UHN',
                #'File1 Value': value1 if pd.notna(value1) else 'Blank in File1',
                #'File2 Value': value2 if pd.notna(value2) else 'Blank in File2'
            #})

        if index in df2.index and index in df1_filtered.index:
            for column in columns_to_compare:
                # Check if the column exists in both DataFrames
                if column in df1_filtered.columns and column in df2.columns:
                    value1 = df1_filtered.at[index, column]
                    value2 = df2.at[index, column]

                    # Normalize "-" in file1 to be equivalent to blank in file2
                    if value1 == "-":
                        value1 = None
                    if pd.isna(value1) and pd.isna(value2):
                        continue  # Both values are NaN, consider them equal
                    if value1 != value2:
                        differences.append({
                            'Physical UHN': index,
                            'Affected Column(s)': column,
                            'SM DATA': value1,
                            'SF DATA': value2
                        })
                else:
                    differences.append({
                        'Physical UHN': index,
                        'Affected Column(s)': column,
                        'SM DATA': df1_filtered.at[index, column] if column in df1_filtered.columns else 'Column not in File1',
                        'SF DATA': df2.at[index, column] if column in df2.columns else 'Column not in File2'
                    })

    # Convert the list of differences into a DataFrame
    differences_df = pd.DataFrame(differences)

    # Save the differences to a new Excel file
    differences_df.to_excel('differences.xlsx', index=False)

def format_excel_file(file):
    """
    Format an Excel file by adjusting column widths, font size, and converting data to a table.

    Args:
        file (str): Path to the Excel file.
    """
    # Load Excel file
    wb = px.load_workbook(file)
    ws = wb.active

    # Convert data to a table
    tab = Table(displayName="Table1", ref=ws.dimensions)
    ws.add_table(tab)

    # Set column width, font size
    for i, column in enumerate(ws.columns):
        max_length = 0
        column_cells = [cell for cell in column]
        column = [str(cell.value) for cell in column]
        if len(column) > 0:
            max_length = max(len(item) for item in column)
        adjusted_width = (max_length + 2) * 2  # adjust this ratio as needed
        if adjusted_width > 35:  # limit column width to 35
            adjusted_width = 35
        ws.column_dimensions[openpyxl.utils.cell.get_column_letter(i+1)].width = adjusted_width

        for cell in column_cells:
            cell.font = Font(size=16)

    # Save the file
    wb.save(file)
    
def compare_uhn(file1, file2, site_to_compare):
    """
    Compare two Excel files and find the differences in a specific column.

    This function loads two Excel files into pandas DataFrames, filters the first DataFrame based on a condition,
    merges the two DataFrames on a specific column, and creates a new DataFrame with the differences found.
    The differences are then saved to a new Excel file, and formatting is applied to the file.

    Returns:
        None
    """

    # Load the two Excel files into pandas DataFrames
    df1 = pd.read_excel(file1)
    df2 = pd.read_excel(file2)

    # Filter the first DataFrame to only include rows where 'Site ID' is 'REDACTED_03'
    df1_filtered = df1[df1['Site ID'] == site_to_compare]

    # Merge the two DataFrames on 'Physical UHN', keeping all entries
    merged = pd.merge(df1_filtered, df2, on='Physical UHN', how='outer', indicator=True)

    # Create a new DataFrame that only includes the rows where the 'Physical UHN' value is only present in one DataFrame
    differences = merged[merged['_merge'] != 'both'][['Physical UHN', '_merge']]

    # Replace 'left_only' and 'right_only' with the respective filenames
    differences['_merge'] = differences['_merge'].cat.rename_categories({'left_only': 'sm_test_data.xlsx', 'right_only': 'sf_test_data.xlsx'})

    # Rename the '_merge' column to 'Differences found in file:'
    differences = differences.rename(columns={'_merge': 'Differences found in file:'})

    # Save the new DataFrame to a new Excel file
    differences.to_excel('uhn_differences.xlsx', index=False)

    # Load the workbook and select the sheet
    wb = load_workbook('uhn_differences.xlsx')
    sheet = wb.active

    # Set the font size and column width
    for column in sheet.columns:
        for cell in column:
            cell.font = cell.font + Font(size=16)
        sheet.column_dimensions[openpyxl.utils.cell.get_column_letter(column[0].column)].width = 50

    # Create a table
    tab = Table(displayName="Table1", ref=sheet.dimensions)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    tab.tableStyleInfo = style
    sheet.add_table(tab)

    # Save the workbook
    wb.save('uhn_differences.xlsx')

# How to use
#file1 = input("Enter the filename of the REDACTED export (e.g. sm_export.xlsx): ") #'sm_test_data.xlsx'
#file2 = input("Enter the filename of the REDACTED export (e.g. sf_export.xlsx): ") #'sf_test_data.xlsx'

def choose_files():
    print("Welcome. This program will compare SM and SF data. It produces two files, one for UHN differences and one for the rest of the data.\n")
    print("First we have to select the [redacted] export file and then the [redacted] export file.\n")
    site_to_compare = input("Finally, enter the site ID to compare (e.g. REDACTED_03): ")
    root = tk.Tk()
    root.withdraw() 
    
    file1 = filedialog.askopenfilename(title="Select the REDACTED export file")
    file2 = filedialog.askopenfilename(title="Select the REDACTED export file")
    
    if file1 and file2:
        compare_excel_files(file1, file2, site_to_compare)
        format_excel_file('differences.xlsx')
        compare_uhn(file1, file2, site_to_compare)
        print("\n\n\nComparison complete. Differences saved to 'differences.xlsx' and 'uhn_differences.xlsx'.")
    else:
        print("No files selected. Exiting...")
        
    
    root.destroy()  # Close the main window

choose_files()
